import cv2
import numpy as np
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import pandas as pd
import os

#This is your excel file location
#If you are changing it paste it below to make sure the syntax is correct first
excel_file_location = "C:\\Users\\Ethan\\OneDrive - Westminster College\\Lignin_Biofilm_Project\\python_image_analysis.xlsx"

#Filtering so that pixels that are outside the range of mean +/- 10*std are set to 0
low_cutoff = float()
hight_cutoff = float()
def filter_image(img, low_cutoff=5, high_cutoff=84):
    mean = mean_intensity(img)
    std = np.std(img)
    low_cutoff = mean - std*10
    hight_cutoff = mean + std*10
    gray_image_filtered = np.copy(gray_image)
    gray_image_filtered[gray_image < low_cutoff] = 0
    gray_image_filtered[gray_image > hight_cutoff] = 0
    return gray_image_filtered

#Blurs image as a another way to decrease noise
def blur_image(img, kernel_size=(5,5), sigma=0):
    blur = cv2.GaussianBlur(img, (5,5), 0)
    return blur

#Figure out %Area of white pixels
def white_area_percentage(img):
    white_pixels = np.sum(img >= 125)
    total_pixels = img.size
    percentage = (white_pixels / total_pixels) * 100
    return percentage


#Median intensity of pixels
def median_intensity(img):
    median_val = np.median(img)
    return median_val

#Mean intensity of pixels
def mean_intensity(img):
    mean_val = np.mean(img)
    return mean_val

#Append results to excel sheet
def append_results(path, row_dict, sheet_name="Sheet1"):
    new_df = pd.DataFrame([row_dict])

    try:
        # Load existing workbook
        book = load_workbook(path)

        # Find the last row in the existing sheet
        startrow = book[sheet_name].max_row

        # Append new row without header
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            new_df.to_excel(writer, sheet_name=sheet_name,
                            startrow=startrow, index=False, header=False)
        file_name = os.path.basename(file)
        print(f"Appended {file_name} to {path}")

    except PermissionError:
        print(f"PermissionError: {path} is locked (Excel or OneDrive may have it open). Close it and try again.")


'''
Start of program execution
'''
#Open file dialog to select a folder
root = tk.Tk()
root.withdraw()  # Hide the root window
#folder that user selects
folder_path = filedialog.askdirectory(title='Select an image folder')
if not folder_path:
    print("No file selected. Exiting.")
    exit()
for file in os.listdir(folder_path):
    #Check if file is a .tif, .tiff
    ext = os.path.splitext(file)[1].lower()  # get file extension
    valid_exts = ['.tif', '.tiff']
    if ext not in valid_exts:
        print(f"Skipping (not an image): {file}")
        continue

    file_path = os.path.join(folder_path, file)
    img = cv2.imread(file_path, cv2.IMREAD_UNCHANGED)

    #Sets image colors to B&W
    gray_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray_image_filtered = filter_image(gray_image, low_cutoff, hight_cutoff)

    blur = blur_image(gray_image_filtered, (5,5), 0)

    #Guassian thresholding method
    thresh_guass = cv2.adaptiveThreshold(
        blur, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,  #You can also try 'cv2.ADAPTIVE_THRESH_MEAN_C' or 'cv2.THRESH_OTSU'
        cv2.THRESH_BINARY,
        725, 0
    )

    percentage = white_area_percentage(thresh_guass)
    median_val = median_intensity(gray_image_filtered)
    mean_val = mean_intensity(gray_image_filtered)

    '''
    Uncomment to see the images
    '''
    #save_path = 'C:/Users/Ethan/Downloads/gray_image_filtered_1.png'
    #cv2.imwrite(save_path, gray_image_filtered)

    print(f'Image: {os.path.basename(file)}\nPercent coverage: {percentage:.2f}% \nMedian: {median_val} \nMean: {mean_val}\nStd: {np.std(gray_image_filtered)}')
    # #Check you have an excel file location then adds the data
    # if excel_file_location is not None:
    #     new_row = {'Image_Name': os.path.basename(file_path), 
    #                'White_Area_Percentage': percentage, 
    #                'Median_Intensity': median_val, 
    #                'Mean_Intensity': mean_val
    #     }
    #     append_results(excel_file_location, new_row, sheet_name='Sheet1')
